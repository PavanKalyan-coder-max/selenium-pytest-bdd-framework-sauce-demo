from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def get_login_data(user_type):

        workbook = load_workbook("testdata/login_data.xlsx")
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if row[0] == user_type:
                return dict(zip(headers, row))

        raise ValueError(f"User '{user_type}' not found in Excel.")